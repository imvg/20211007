# -*- coding: utf-8 -*-
import logging
import pymysql
from openpyxl import load_workbook, Workbook
import time
import os

logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)


connection = pymysql.connect(host="hk-cdb-aumsfh1d.sql.tencentcdb.com", user="root", password="wvqhthpbd!Wv$da4", port=63734, db="db_yy")


def getData(start, aid=False):
    try:
        with connection.cursor() as cur:
            stop = start + 86400
            if not aid:
                sql = f"select dev_info from t_user where os='android' and regtime >= {start} and regtime <= {stop};"
            else:
                sql = f"select dev_info from t_user where os='android' and aid={aid} and regtime >= {start} and regtime <= {stop};"
            cur.execute(sql)
            res = cur.fetchall()
            logging.info("机型获取完成")
            if not os.path.exists('统计.xlsx'):
                wb = Workbook()
            else:
                wb = load_workbook('统计.xlsx')
            pinpai = {}
            version = {}
            ws = wb.active
            daytime = time.strftime("%Y-%m-%d", time.localtime(start + (4 * 3600)))
            ws1 = wb.create_sheet(daytime)
            ws1["A1"] = '手机品牌'
            ws1["B1"] = '数量'
            ws1["C1"] = '占比'
            ws1["E1"] = '系统版本'
            ws1["F1"] = '数量'
            ws1["G1"] = '占比'
            logging.info(f"生成机型库")
            for data in res:
                device = eval(data[0])
                pinpai[f"{device['brand'].lower()}"] = 0
                version[f"{device['version'].lower()}"] = 0
            logging.info(f"查询各品牌上报数据")
            for data in pinpai.keys():
                if not aid:
                    countSql = f"select count(id) from t_character where params like '%{data.lower()}%' and ctime>={start} and ctime <={stop};"
                else:
                    countSql = f"select count(id) from t_character where aid={aid} and params like '%{data.lower()}%' and ctime>={start} and ctime <={stop};"
                cur.execute(countSql)
                countRes = cur.fetchone()[0]
                logging.info(f"匹配品牌数据 {countSql} RETURN {countRes}")
                pinpai[data.lower()] = countRes
            logging.info(f"查询各版本上报数据")
            for data in version.keys():
                if not aid:
                    countSql = f"select count(id) from t_character where params like '%android {data.lower()}%' and ctime>={start} and ctime <={stop};"
                else:
                    countSql = f"select count(id) from t_character where aid={aid} and params like '%android {data.lower()}%' and ctime>={start} and ctime <={stop};"
                cur.execute(countSql)
                countRes = cur.fetchone()[0]
                logging.info(f"匹配版本数据 {countSql} RETURN {countRes}")
                version[data.lower()] = countRes
            logging.info(version)
            logging.info(f"数据匹配完成")
            pinpai_col = 2
            version_col = 2
            pinpai_total = sum(pinpai.values())
            version_total = sum(version.values())
            pinpai = sorted(pinpai.items(), key=lambda x: x[1], reverse=True)
            version = sorted(version.items(), key=lambda x: x[1], reverse=True)
            logging.info(f"写入Excel")
            for d in pinpai:
                ws1[f"A{pinpai_col}"] = d[0]
                ws1[f"B{pinpai_col}"] = d[1]
                ws1[f"C{pinpai_col}"] = f"{'%.2f' % (int(d[1]) / pinpai_total * 100)}%"
                pinpai_col += 1
            for d in version:
                ws1[f"E{version_col}"] = d[0]
                ws1[f"F{version_col}"] = d[1]
                ws1[f"G{version_col}"] = f"{'%.2f' % (int(d[1]) / version_total * 100)}%"
                version_col += 1
            wb.save('统计.xlsx')
            wb.close()
    except Exception as e:
        logging.error(e)


startData = 1627574400
for d in range(5):
    logging.info(f'统计日期 {time.strftime("%Y-%m-%d", time.localtime(startData + (4 * 3600)))}')
    getData(startData, aid=265)
    startData += 86400

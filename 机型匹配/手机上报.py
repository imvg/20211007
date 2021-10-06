# -*- coding: utf-8 -*-
import logging
import pymysql
from openpyxl import load_workbook, Workbook
import time
import os

logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)


connection = pymysql.connect(host="hk-cdb-aumsfh1d.sql.tencentcdb.com", user="root", password="wvqhthpbd!Wv$da4", port=63734, db="db_yy")


def getData(start, aid=False):
    try:
        with connection.cursor() as cur:
            stop = start + 86400
            if not aid:
                sql = f"select dev_info from t_user where os='android' and regtime >= {start} and regtime <= {stop};"
            else:
                sql = f"select dev_info from t_user where os='android' and aid={aid} and regtime >= {start} and regtime <= {stop};"
            cur.execute(sql)
            res = cur.fetchall()
            logging.info(res)
            if not os.path.exists('统计.xlsx'):
                wb = Workbook()
            else:
                wb = load_workbook('统计.xlsx')
            ws = wb.active
            daytime = time.strftime("%Y-%m-%d", time.localtime(start + (4 * 3600)))
            ws1 = wb.create_sheet(daytime)
            pinpai = {}
            version = {}
            model = {}
            ws1["A1"] = '手机品牌'
            ws1["B1"] = '数量'
            ws1["C1"] = '占比'
            ws1["E1"] = '系统版本'
            ws1["F1"] = '数量'
            ws1["G1"] = '占比'
            ws1["I1"] = '品牌机型'
            ws1["J1"] = '数量'
            ws1["K1"] = '占比'
            for data in res:
                device = eval(data[0])
                logging.info(device)
                if device['brand'].lower() in pinpai.keys():
                    pinpai[f"{device['brand'].lower()}"] = pinpai[f"{device['brand'].lower()}"] + 1
                else:
                    pinpai[f"{device['brand'].lower()}"] = 1
                if device['version'].lower() in version.keys():
                    version[f"{device['version'].lower()}"] = version[f"{device['version'].lower()}"] + 1
                else:
                    version[f"{device['version'].lower()}"] = 1
                if device['model'].lower() in model.keys():
                    model[f"{device['model'].lower()}"] = model[f"{device['model'].lower()}"] + 1
                else:
                    model[f"{device['model'].lower()}"] = 1
            pinpai_col = 1
            version_col = 1
            model_col = 1
            pinpai_total = sum(pinpai.values())
            version_total = sum(version.values())
            model_total = sum(model.values())
            pinpai = sorted(pinpai.items(), key=lambda x: x[1], reverse=True)
            version = sorted(version.items(), key=lambda x: x[1], reverse=True)
            model = sorted(model.items(), key=lambda x: x[1], reverse=True)
            for d in pinpai:
                ws1[f"A{pinpai_col+1}"] = d[0]
                ws1[f"B{pinpai_col+1}"] = d[1]
                ws1[f"C{pinpai_col+1}"] = f"{'%.2f' % (int(d[1]) / pinpai_total * 100)}%"
                pinpai_col += 1
            for d in version:
                ws1[f"E{version_col + 1}"] = d[0]
                ws1[f"F{version_col + 1}"] = d[1]
                ws1[f"G{version_col + 1}"] = f"{'%.2f' % (int(d[1]) / version_total * 100)}%"
                version_col += 1
            for d in model:
                ws1[f"I{model_col + 1}"] = d[0]
                ws1[f"J{model_col + 1}"] = d[1]
                ws1[f"K{model_col + 1}"] = f"{'%.2f' % (int(d[1]) / model_total * 100)}%"
                model_col += 1
            wb.save('统计.xlsx')
            wb.close()
    except Exception as e:
        logging.error(e)


startData = 1627574400
for d in range(5):
    logging.info(f'统计日期 {time.strftime("%Y-%m-%d", time.localtime(startData + (4 * 3600)))}')
    getData(startData, aid=265)
    startData += 86400


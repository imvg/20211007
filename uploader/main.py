# -*- coding: utf-8 -*-
import logging
from tkinter import Tk
from tkinter import END
from tkinter import Label
from tkinter import LabelFrame
from tkinter import StringVar
from tkinter import IntVar
from tkinter import Radiobutton
from tkinter import filedialog
from tkinter import scrolledtext
from tkinter import ttk
from tkinter import messagebox
from qcloud_cos import CosConfig
from qcloud_cos import CosS3Client
from openpyxl import load_workbook
import platform
import pymysql
import threading
import time
import subprocess
import os
import cv2
import requests
import re


class App:
    def __init__(self):
        self.root = Tk()
        self.root.title("撸呗 - 视频上传工具 v2.1")
        self.root.maxsize(540, 600)
        self.root.minsize(540, 600)

        book = ttk.Notebook(self.root)
        self.note1 = ttk.Frame(book)

        self.setFrame = LabelFrame(self.note1, text="视频信息", height=380, width=500)
        self.setZone()
        self.setFrame.place(x=20, y=0)

        self.logFrame = LabelFrame(self.note1, text="处理日志", height=125, width=500)
        self.logZone()
        self.logFrame.place(x=20, y=390)

        self.note2 = ttk.Frame(book)

        self.note2ConfigSet = LabelFrame(self.note2, text="视频信息", height=200, width=500)
        self.note2Config()
        self.note2ConfigSet.place(x=20, y=0)

        self.note2LogSet = LabelFrame(self.note2, text="处理日志", height=320, width=500)
        self.note2Log()
        self.note2LogSet.place(x=20, y=210)

        book.add(self.note1, text="单个上传")
        book.add(self.note2, text="批量上传")

        book.pack(expand=1, fill="both")

        self.setlog('准备就绪!')
        self.root.mainloop()

    def queren1(self):
        askback = messagebox.askokcancel('温馨提示', '确认配置无误，继续执行？')
        if askback == True:
            self.upload()

    def queren2(self):
        askback = messagebox.askokcancel('温馨提示', '确认配置无误，继续执行？')
        if askback == True:
            self.upload2()

    def setZone(self):
        Label(self.setFrame, text='封面文件').place(x=80, y=20)
        self.cover_en = ttk.Entry(self.setFrame, width=25)
        self.cover_en.config(state='disabled')
        self.cover_en.place(x=145, y=20)
        self.cover_bu = ttk.Button(self.setFrame, text='选择', command=self.getcover)
        self.cover_bu.place(x=345, y=19)

        Label(self.setFrame, text='视频文件').place(x=80, y=50)
        self.vod_en = ttk.Entry(self.setFrame, width=25)
        self.vod_en.config(state='disabled')
        self.vod_en.place(x=145, y=50)
        self.vod_bu = ttk.Button(self.setFrame, text='选择', command=self.getvod)
        self.vod_bu.place(x=345, y=49)

        Label(self.setFrame, text='视频标题').place(x=80, y=80)
        self.title_en = ttk.Entry(self.setFrame, width=20)
        self.title_en.insert(END, "测试视频")
        self.title_en.place(x=145, y=79)

        Label(self.setFrame, text='视频番号').place(x=80, y=110)
        self.fanhao_en = ttk.Entry(self.setFrame, width=20)
        self.fanhao_en.insert(END, "666888")
        self.fanhao_en.place(x=145, y=109)

        Label(self.setFrame, text='女优原名').place(x=80, y=140)
        self.actor_realname_en = ttk.Entry(self.setFrame, width=20)
        self.actor_realname_en.insert(END, "李狗蛋")
        self.actor_realname_en.place(x=145, y=139)

        Label(self.setFrame, text='女优中文名').place(x=70, y=170)
        self.actor_chinaname_en = ttk.Entry(self.setFrame, width=20)
        self.actor_chinaname_en.insert(END, "李狗蛋")
        self.actor_chinaname_en.place(x=145, y=169)

        Label(self.setFrame, text='视频分类').place(x=80, y=200)
        self.video_type_en = ttk.Entry(self.setFrame, width=20)
        self.video_type_en.insert(END, "日韩")
        self.video_type_en.place(x=145, y=199)

        Label(self.setFrame, text='存储位置').place(x=80, y=230)
        self.file_save_en = ttk.Entry(self.setFrame, width=20)
        self.file_save_en.insert(END, "movie/cn/MD")
        self.file_save_en.place(x=145, y=229)

        self.vodtype = StringVar()
        self.have = Radiobutton(self.setFrame, variable=self.vodtype, value="1", text='有码')
        self.nohave = Radiobutton(self.setFrame, variable=self.vodtype, value="2", text='无码')
        self.unknow = Radiobutton(self.setFrame, variable=self.vodtype, value="0", text='未知')
        self.vodtype.set("0")
        self.have.place(x=130, y=270)
        self.nohave.place(x=230, y=270)
        self.unknow.place(x=330, y=270)

        self.upload_bu = ttk.Button(self.setFrame, text='上传', command=self.queren1)
        self.upload_bu.place(x=220, y=320)

    def logZone(self):
        self.log = scrolledtext.ScrolledText(self.logFrame, height=7, width=66, state='disabled')
        self.log.place(x=10, y=0)

    def setlog(self, log):
        self.log.config(state='normal')
        self.log.insert(END, f'{time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())}: {log} \n')
        self.log.see(END)
        self.log.config(state='disabled')

        self.log2.config(state='normal')
        self.log2.insert(END, f'{time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())}: {log} \n')
        self.log2.see(END)
        self.log2.config(state='disabled')

    def getcover(self):
        self.filePath = filedialog.askopenfilename(title=u'选择文件',
                                                   filetypes=[("图片文件", "*.jpg"), ("图片文件", "*.jpeg"), ("图片文件", "*.png"), ("图片文件", "*.gif")])
        self.cover_en.config(state='normal')
        self.cover_en.delete(0, "end")
        self.cover_en.insert(0, self.filePath)
        self.cover_en.config(state='disabled')

    def getvod(self):
        self.filePath = filedialog.askopenfilename(title=u'选择文件', filetypes=[("视频文件", "*.mp4"), ("视频文件", "*.mov"), ("视频文件", "*.avi")])
        self.vod_en.config(state='normal')
        self.vod_en.delete(0, "end")
        self.vod_en.insert(0, self.filePath)
        self.vod_en.config(state='disabled')

    def getConfig(self):
        configPath = filedialog.askopenfilename(title=u'选择文件', filetypes=[("Excel", "*.xlsx")])
        self.config_en.config(state='normal')
        self.config_en.delete(0, "end")
        self.config_en.insert(0, configPath)
        self.config_en.config(state='disabled')

    def note2Config(self):
        Label(self.note2ConfigSet, text='配置文件').place(x=80, y=20)
        self.config_en = ttk.Entry(self.note2ConfigSet, width=25)
        self.config_en.config(state='disabled')
        self.config_en.place(x=145, y=20)

        self.config_bu = ttk.Button(self.note2ConfigSet, text='选择', command=self.getConfig)
        self.config_bu.place(x=345, y=19)

        self.imageFirst = IntVar()
        self.imageFirst.set(1)
        self.imageFirst_bu = ttk.Checkbutton(self.note2ConfigSet, text='封面取视频第一帧', onvalue=0, offvalue=1, variable=self.imageFirst)
        self.imageFirst_bu.place(x=200, y=60)

        self.note2Upload_bu = ttk.Button(self.note2ConfigSet, text="上传", command=self.queren2)
        self.note2Upload_bu.place(x=220, y=120)

    def note2Log(self):
        self.log2 = scrolledtext.ScrolledText(self.note2LogSet, height=22, width=66, state='disabled')
        self.log2.place(x=10, y=0)

    def lock(self):
        self.cover_bu.config(state='disabled')
        self.vod_bu.config(state='disabled')
        self.title_en.config(state='disabled')
        self.fanhao_en.config(state='disabled')
        self.have.config(state='disabled')
        self.nohave.config(state='disabled')
        self.unknow.config(state='disabled')
        self.actor_realname_en.config(state='disabled')
        self.actor_chinaname_en.config(state='disabled')
        self.video_type_en.config(state='disabled')
        self.file_save_en.config(state='disabled')
        self.upload_bu.config(state='disabled')
        self.config_bu.config(state='disabled')
        self.note2Upload_bu.config(state='disabled')
        self.imageFirst_bu.config(state='disabled')

    def unlock(self):
        self.cover_bu.config(state='normal')
        self.vod_bu.config(state='normal')
        self.title_en.config(state='normal')
        self.fanhao_en.config(state='normal')
        self.have.config(state='normal')
        self.nohave.config(state='normal')
        self.unknow.config(state='normal')
        self.actor_realname_en.config(state='normal')
        self.actor_chinaname_en.config(state='normal')
        self.video_type_en.config(state='normal')
        self.file_save_en.config(state='normal')
        self.upload_bu.config(state='normal')
        self.config_bu.config(state='normal')
        self.note2Upload_bu.config(state='normal')
        self.imageFirst_bu.config(state='normal')

    def upload(self):
        threading.Thread(target=self.action).start()

    def upload2(self):
        threading.Thread(target=self.action2).start()

    def action(self):
        if self.vod_en.get() and self.title_en.get() and self.fanhao_en.get() and self.file_save_en.get() and self.fanhao_en.get():
            self.lock()
            self.log.config(state='normal')
            self.log.insert(END, '=' * 60 + '\n')
            self.log.see(END)
            self.log.config(state='disabled')
            self.vodname = re.split('/', self.vod_en.get())[-1]

            bucket_url = 'https://sourcefile-1304080031.cos.accelerate.myqcloud.com'
            existed = self.checkExisted(
                f"{bucket_url}/{self.file_save_en.get()}/{self.fanhao_en.get()}/{self.fanhao_en.get()}.m3u8",
                self.vodname, self.fanhao_en.get())
            if existed == True:
                self.unlock()
                return False
            self.vodduration = self.getVideoTime(self.vod_en.get(), self.vodname)
            if self.vodduration == False:
                self.unlock()
                return False
            self.vodreso = self.getVideoResolution(self.vod_en.get(), self.vodname)
            if self.vodreso == False:
                self.unlock()
                return False
            if self.imageFirst.get() == 0:
                self.setlog(f"{self.vodname} 正在取视频第一帧...")
                firstRes = self.getFirstImage(self.vod_en.get(), f"tmp/real-{self.fanhao_en.get()}.jpg")
                self.fanhao_en.config(state='normal')
                self.fanhao_en.delete(0, END)
                self.fanhao_en.insert(END, f"tmp/real-{self.fanhao_en.get()}.jpg")
                self.fanhao_en.config(state='disabled')
                if not firstRes:
                    self.setlog(f"{self.vodname} 取第一帧失败!")
                    return False
            if self.encryptImage(self.cover_en.get(), self.vodname, self.fanhao_en.get()) == False:
                self.unlock()
                return False
            if self.vodSplit(self.vod_en.get(), self.vodname, self.fanhao_en.get()) == False:
                self.unlock()
                return False
            previewPath = f"tmp/preview-{int(time.time())}.m3u8"
            if self.makePreview(f"tmp/{self.fanhao_en.get()}.m3u8", previewPath, self.vodname) == False:
                return False
            if self.uploadVideo(self.vodname, self.file_save_en.get(), self.fanhao_en.get(), self.cover_en.get(), previewPath) == False:
                self.unlock()
                return False
            data = {
                "vodname": self.vodname,
                "cover": self.cover_en.get(),
                "fanhao": self.fanhao_en.get(),
                "title": self.title_en.get(),
                "name": self.actor_realname_en.get(),
                "chinaname": self.actor_chinaname_en.get(),
                "videocode": self.vodtype.get(),
                "videotype": self.video_type_en.get(),
                "file_save": self.file_save_en.get(),
                "vodreso": self.vodreso,
                "vodduration": self.vodduration,
                "preview": re.split('/', previewPath)[-1],
            }
            if self.saveBase(data) == False:
                self.unlock()
                return False
            self.setlog("任务完成!")
            self.unlock()
        else:
            self.setlog("输入的视频信息不完整")

    def action2(self):
        try:
            self.lock()
            if self.config_en.get() and os.path.exists(self.config_en.get()):
                table = load_workbook(self.config_en.get())
                sh = table[table.sheetnames[0]]
                if sh.max_column == 9:
                    for index in range(2, int(sh.max_row) + 1):
                        fanhao = sh[f"A{index}"].value
                        self.setlog("=" * 10 + f" {fanhao} {index - 1}/{int(sh.max_row) - 1} " + "=" * 10)
                        data = {"fanhao": sh[f"A{index}"].value,
                                "title": sh[f"B{index}"].value,
                                "name": sh[f"C{index}"].value,
                                "chinaname": sh[f"D{index}"].value,
                                "videotype": sh[f"E{index}"].value,
                                "youwuma": sh[f"F{index}"].value,
                                "savepath": sh[f"G{index}"].value,
                                "cover": str(sh[f"H{index}"].value).replace("/", "\\"),
                                "movie": str(sh[f"I{index}"].value).replace("/", "\\")
                                }
                        self.work(data)
                table.close()
                self.setlog("任务完成!")
                self.unlock()
            else:
                self.setlog(f"配置文件异常, 请检查!")
                self.unlock()
        except Exception as e:
            self.setlog(f"处理异常 {e}")

    def work(self, data):
        try:
            bucket_url = 'https://sourcefile-1304080031.cos.accelerate.myqcloud.com'
            vodname = re.split("\\\\", data['movie'])[-1]
            existed = self.checkExisted(f"{bucket_url}/{data['savepath']}/{data['fanhao']}/{data['fanhao']}.m3u8", vodname, data['fanhao'])
            if existed:
                return False
            vodduration = self.getVideoTime(data['movie'], vodname)
            if vodduration == False:
                return False
            vodreso = self.getVideoResolution(data['movie'], vodname)
            if vodreso == False:
                return False
            if self.imageFirst.get() == 0:
                self.setlog(f"{vodname} 正在取视频第一帧...")
                firstRes = self.getFirstImage(data['movie'], f"tmp/real-{data['fanhao']}.jpg")
                data['cover'] = f"tmp/real-{data['fanhao']}.jpg"
                if not firstRes:
                    self.setlog(f"{vodname} 取第一帧失败!")
                    return False
            if self.encryptImage(data['cover'], vodname, data['fanhao']) == False:
                return False
            if self.vodSplit(data['movie'], vodname, data['fanhao']) == False:
                return False

            previewPath = f"tmp/preview-{int(time.time())}.m3u8"
            if self.makePreview(f"tmp/{data['fanhao']}.m3u8", previewPath, vodname) == False:
                return False

            if self.uploadVideo(vodname, data['savepath'], data['fanhao'], data['cover'], previewPath) == False:
                return False
            data = {
                "vodname": vodname,
                "cover": data['cover'],
                "fanhao": data['fanhao'],
                "title": data['title'],
                "name": data['name'],
                "chinaname": data['chinaname'],
                "videocode": data['youwuma'],
                "videotype": data['videotype'],
                "file_save": data['savepath'],
                "vodreso": vodreso,
                "vodduration": vodduration,
                "preview": re.split('/', previewPath)[-1]
            }
            if self.saveBase(data) == False:
                return False
        except Exception as e:
            self.setlog(f"处理异常 {e}")

    def vodSplit(self, vodpath, vodname, fanhao):
        self.setlog(f"{vodname} 视频切片加密中...")
        spres = True
        try:
            if not os.path.isdir('tmp'):
                os.mkdir('tmp')
            if os.path.exists(".\\ffmpeg-win.exe"):
                command = f".\\ffmpeg-win.exe -y -i {vodpath} -force_key_frames 'expr:gte(t,n_forced*1)' -hls_time 2 -strict -2 -c:a aac -c:v libx264 -hls_time 2 -hls_list_size 0 -hls_segment_filename tmp/{fanhao}-%05d.ts -hls_key_info_file enc.keyinfo tmp/{fanhao}.m3u8"
            else:
                self.setlog("ffmpeg程序文件不存在")
                return False
            ret = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
            while True:
                if ret.poll() is None:
                    # 切片详细日志
                    runlog = ret.stdout.readline()
                    continue
                else:
                    if ret.wait() != 0:
                        self.setlog('视频切片失败999!')
                        spres = False
                        break
                    else:
                        break
        except Exception as e:
            self.setlog(f'视频切片失败! {e}')
            spres = False
        return spres

    def encryptImage(self, imgpath, vodname, fanhao):
        try:
            self.setlog(f"{vodname} 加密视频封面图...")
            data = open(imgpath, 'rb').read()
            key = os.urandom(8)
            type = re.split('\.', imgpath)[-1]
            with open(f'tmp/{fanhao}.{type}', 'wb') as enc:
                enc.write(key)
                enc.write(data)
            if re.split('/', imgpath)[-1].startswith('real-'):
                os.remove(imgpath)
            return True
        except Exception as e:
            self.setlog(f'封面加密失败! {e}')
            return False

    def getVideoTime(self, vodpath, vodname):
        self.setlog(f"{vodname} 获取视频时长...")
        if not os.path.exists(vodpath):
            self.setlog(vodpath)
            self.setlog(f"{vodname} 视频文件不存在")
            return False
        try:
            cap = cv2.VideoCapture(vodpath)
            if cap.isOpened():
                rate = cap.get(5)
                FrameNumber = cap.get(7)
                duration = FrameNumber / rate * 1000
                return int(duration)
            else:
                self.setlog(f"{vodname} 视频文件异常, 无法获取时长")
                return False
        except Exception as e:
            self.setlog(f"{vodname} 视频文件异常, 无法获取时长 {e}")
            self.setlog(e)
            return False

    def getVideoResolution(self, vodpath, vodname):
        self.setlog(f"{vodname} 获取视频分辨率...")
        if not os.path.exists(vodpath):
            self.setlog(f"{vodname} 视频文件不存在")
            return False
        try:
            cap = cv2.VideoCapture(vodpath)
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            return f"{width}x{height}"
        except Exception as e:
            self.setlog("视频分辨率获取失败")
            self.setlog(e)
            return False

    def checkExisted(self, url, vodname, fanhao):
        self.setlog(f"{vodname} 检查视频是否存在...")
        res = False
        try:
            connection = pymysql.connect(host="hk-cdb-aumsfh1d.sql.tencentcdb.com", user="root",
                                         password="wvqhthpbd!Wv$da4", port=63734, db="db_yy", connect_timeout=10,
                                         write_timeout=10)
        except Exception as e:
            self.setlog(f"连接失败 {e}")
        else:
            try:
                with connection.cursor() as cur:
                    sql1 = f"select * from t_video where vnum='{fanhao}' and `status`>-2"
                    cur.execute(sql1)
                    if len(cur.fetchall()) >= 1:
                        self.setlog(f"{vodname} 视频表存在该番号的视频")
                        res = True
                    sql2 = f"select * from t_video_stock where vnum='{fanhao}' and `status`>0"
                    cur.execute(sql2)
                    if len(cur.fetchall()) >= 1:
                        self.setlog(f"{vodname} 库存表存在该番号的视频")
                        res = True
                    ret = requests.get(url=url, timeout=10)
                    if ret.status_code == 200:
                        self.setlog(f"{vodname} 视频文件已存在，请勿重复上传.")
                        res = True
            except Exception as e:
                self.setlog(f"番号校验失败 {e}")
                res = True
            finally:
                connection.close()
        finally:
            return res

    def uploadVideo(self, vodname, savepath, fanhao, cover, preview):
        self.setlog(f"{vodname} 视频上传处理中...")
        secret_id = ''
        secret_key = ''
        region = 'ap-hongkong'
        domain = 'sourcefile-1304080031.cos.accelerate.myqcloud.com'
        client = CosS3Client(CosConfig(Region=region, SecretId=secret_id, SecretKey=secret_key, Domain=domain))
        cos_bucket = 'sourcefile-1304080031'
        dir = savepath.strip('/').replace('\\', '/')

        def uploadfile(sfile, dfile):
            try:
                client.upload_file(
                    Bucket=cos_bucket,
                    LocalFilePath=sfile,
                    Key=dfile,
                    PartSize=1,
                    MAXThread=10,
                    EnableMD5=False
                )
                os.remove(sfile)
                return True
            except Exception as e:
                self.setlog(f"文件上传异常 {e}")
                return False

        m3u8 = open(f'tmp/{fanhao}.m3u8', 'r')
        for ts in m3u8.readlines():
            if not ts.startswith('#'):
                if uploadfile(f'tmp/{ts.strip()}', dir + f"/{fanhao}/{ts.strip()}") is not True:
                    return False
        m3u8.close()
        if uploadfile(f'tmp/{fanhao}.m3u8', dir + f"/{fanhao}/{fanhao}.m3u8") is not True:
            return False
        imgtype = re.split('\.', cover)[-1]
        if uploadfile(f'tmp/{fanhao}.{imgtype}', dir + f'/{fanhao}/{fanhao}.{imgtype}') is not True:
            return False
        previewFile = re.split('/', preview)[-1]
        if uploadfile(preview, dir + f'/{fanhao}/{previewFile}') is not True:
            return False
        return True

    def getFirstImage(self, mp4path, imagepath):
        try:
            cap = cv2.VideoCapture(mp4path)
            cap.set(cv2.CAP_PROP_POS_FRAMES, 1)
            if cap.isOpened():
                rval, frame = cap.read()
                cv2.imencode('.jpg', frame)[1].tofile(imagepath)
                return True
            else:
                return False
        except Exception as e:
            self.setlog(f"取第一帧失败 {e}")
            return False

    def getPreviewData(self, dataFile, start, end):
        try:
            res = ""
            line = 0
            totalTime = 0
            startLine = 0
            console = False
            console2 = False
            console3 = False
            data = open(dataFile, 'r')
            filedata = data.readlines()
            for ts in filedata:
                line += 1
                if console:
                    res += ts
                if 'EXTINF' in ts:
                    if not console3:
                        startLine = line
                        console3 = True
                    longtime = float(re.split(':', ts)[-1].replace(',', ''))
                    totalTime += longtime
                    if totalTime >= start and not console2:
                        res += ts
                        console = True
                        console2 = True
                    if totalTime >= end and console:
                        res += filedata[line]
                        console = False
            titleData = ""
            for title in filedata[:startLine - 1]:
                titleData += title
            res = titleData + res
            if filedata[-1] not in res:
                res += filedata[-1]
            data.close()
            return res
        except Exception as e:
            self.setlog(f"预览版生成失败 {e}")
            return False

    def makePreview(self, m3u8Path, dstPath, vodname):
        try:
            self.setlog(f"{vodname} 正在生成300秒预览版...")
            previewData = self.getPreviewData(m3u8Path, 1, 300)
            if previewData:
                with open(dstPath, 'w') as f:
                    f.write(previewData)
                return True
            else:
                return False
        except Exception as e:
            self.setlog(e)
            return False

    def saveBase(self, data):
        # 打开数据库连接
        self.setlog(f"{data['vodname']} 视频入库中...")
        try:
            connection = pymysql.connect(host="hk-cdb-aumsfh1d.sql.tencentcdb.com", user="root", password="wvqhthpbd!Wv$da4", port=63734, db="db_yy", connect_timeout=10, write_timeout=10)
            # connection = pymysql.connect(host="101.32.202.66", user="root", password="1q2w3e4r..A", port=3306, db="db_yy_1", connect_timeout=10, write_timeout=10)
        except Exception as e:
            self.setlog("数据库连接失败")
            self.setlog(e)
            return False
        try:
            title = data['title']
            cover_type = re.split('\.', data['cover'])[-1]
            cover_uri = "/" + data['file_save'] + "/" + data['fanhao'] + "/" + data['fanhao'] + "." + cover_type
            vod_url = "/" + data['file_save'] + "/" + data['fanhao'] + "/" + data['fanhao'] + ".m3u8"
            district = data['videotype']
            vnum = str(data['fanhao'])
            status = "1"
            videotype = "1"
            videocode = data['videocode']
            ctime = int(time.time())
            resolution = data['vodreso']
            duration = data['vodduration']
            name = data['name']
            ch_name = data['chinaname']
            preview = "/" + data['file_save'] + "/" + data['fanhao'] + "/" + data['preview']
            with connection.cursor() as cursor:
                sql = f"insert into t_video_stock(`title`,`cover_uri`,`video_uri`,`video_uri1`,`video_uri2`,`video_uri3`,`district`,`vnum`,`video_type`,`video_code`,`description`,`resolution`,`duration`,`md5`,`name`,`ch_name`,`ctime`,`status`)"\
                                        f" value('{title}','{cover_uri}','{vod_url}','{vod_url}','{vod_url}','{preview}','{district}','{vnum}','{videotype}','{videocode}','','{resolution}','{duration}','','{name}','{ch_name}','{ctime}','{status}')"
                cursor.execute(sql)
                connection.commit()
        except Exception as e:
            self.setlog("视频入库失败...")
            self.setlog(e)
            connection.close()
            return False
        finally:
            connection.close()
            return True


if __name__ == '__main__':
    App()
